import openpyxl 

wb= openpyxl.load_workbook("example.zlsx") 
print (wb.sheetnames) 
# wb.active selects the first available sheet (sheet 1) 
sheet = wb.active 
print (sheet.title) 
for i in range(1, sheet.max_row+1): 
      # print all the data in column A,B,C 
      # ljust() is used to make it easier to read 
     Print(i , str(sheet.cell(row=i , column=1).value).ljust(25, ‘ ‘), 
     Sheet.cell(row=i, column=2).value.ljust(15,’ ‘), 
     Str(sheet.cell(row=i, column=3).value)) 